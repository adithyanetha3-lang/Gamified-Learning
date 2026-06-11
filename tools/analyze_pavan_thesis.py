from pathlib import Path
from collections import Counter
import json
import math
import re

import openpyxl


XLSX = Path(r"C:/Users/adith/AppData/Local/Packages/5319275A.WhatsAppDesktop_cv1g1gvanyjgm/LocalState/sessions/F53961CCE067D31894D85901B81F73D5755EAA99/transfers/2026-23/PavanThesis.0.xlsx")
OUT = Path("outputs/pavan_mandamarri/analysis.json")


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value.strip())
    return value


def rows_for(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [clean(v) or f"Column {i+1}" for i, v in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        if any(clean(v) not in (None, "") for v in row):
            data.append({headers[i]: clean(row[i]) if i < len(row) else None for i in range(len(headers))})
    return headers, data


def freq(data, key, limit=12):
    counts = Counter()
    for row in data:
        val = clean(row.get(key))
        if val not in (None, ""):
            counts[str(val)] += 1
    total = sum(counts.values()) or 1
    return [{"value": k, "count": v, "percent": round(v * 100 / total, 1)} for k, v in counts.most_common(limit)]


def numeric_values(data, key):
    vals = []
    for row in data:
        v = row.get(key)
        if isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v):
            vals.append(float(v))
        elif isinstance(v, str):
            nums = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", v)
            if nums:
                try:
                    vals.append(float(nums[0].replace(",", "")))
                except ValueError:
                    pass
    return vals


def stats(data, key):
    vals = numeric_values(data, key)
    if not vals:
        return None
    vals_sorted = sorted(vals)
    n = len(vals_sorted)
    return {
        "count": n,
        "min": round(vals_sorted[0], 2),
        "max": round(vals_sorted[-1], 2),
        "mean": round(sum(vals_sorted) / n, 2),
        "median": round(vals_sorted[n // 2] if n % 2 else (vals_sorted[n // 2 - 1] + vals_sorted[n // 2]) / 2, 2),
    }


def yes_count(data, key):
    yes = 0
    no = 0
    other = 0
    for row in data:
        val = str(clean(row.get(key)) or "").lower()
        if val.startswith("yes"):
            yes += 1
        elif val.startswith("no"):
            no += 1
        elif val:
            other += 1
    total = yes + no + other or 1
    return {"yes": yes, "no": no, "other": other, "yes_percent": round(yes * 100 / total, 1), "n": total}


def maybe_index_sheet(headers):
    joined = " ".join(map(str, headers)).lower()
    return any(term in joined for term in ["cvi", "vulnerability", "sensitivity index", "exposure index", "adaptive"])


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheets = {}
    for ws in wb.worksheets:
        headers, data = rows_for(ws)
        sheets[ws.title] = {"headers": headers, "data": data}

    basic = sheets["Basic Information"]["data"]
    edu = sheets["Section_B_Education_Health"]["data"]
    liv = sheets["Section_C_Livelihood"]["data"]
    exp = sheets["Section_D_Exposure"]["data"]
    sens = sheets[" Sensitivity (socio-economic v"]["data"]
    adap = sheets["Adaptive Capacity"]["data"]

    summary = {
        "sample_size": len(basic),
        "sheet_names": wb.sheetnames,
        "demographics": {
            "village": freq(basic, "Village"),
            "gender": freq(basic, "Gender"),
            "religion": freq(basic, "Religion"),
            "caste": freq(basic, "Caste"),
            "house_type": freq(basic, "HH_type"),
            "age": stats(basic, "Age"),
            "household_size": stats(basic, "HH Size"),
        },
        "education_health": {
            "education": freq(edu, "B1_Education"),
            "vocational_training": yes_count(edu, "B2_Vocational Training \\Techincal Training"),
            "chronic_illness": yes_count(edu, "B3_Chronic_Illness"),
            "chronic_illness_frequency": freq(edu, "B3_Chronic_Illness"),
            "affects_earning": yes_count(edu, "B3b_Affect_Earning"),
            "hospitalized": yes_count(edu, "B4_Hospitalized"),
            "medical_expense": yes_count(edu, "B4_Medical Expense"),
            "health_insurance": yes_count(edu, "B5_Health_Insurance"),
        },
        "livelihood": {
            "main_livelihood": freq(liv, "C1_Main_Livelihood"),
            "active_members": stats(liv, "C2_Economic_Active members"),
            "currently_mining": yes_count(liv, "C3_Mining (Currently)"),
            "past_mining": yes_count(liv, "C4_Mining(Past)"),
            "income_source": freq(liv, "C5_Main sources of income"),
            "income": stats(liv, "C6_Income Category"),
            "income_stability": freq(liv, "C7_Income Stability"),
            "employment_type": freq(liv, "C8_Mining Employment Type"),
            "formal_documentation": yes_count(liv, "C9_Formal Documentation"),
        },
        "environmental_exposure": {
            "distance_mine": freq(exp, "D1_Distance_Mine"),
            "dust": yes_count(exp, "D2_Dust or airborne pollution from mining"),
            "noise": freq(exp, "D3_ Noise disturbance"),
            "accident": yes_count(exp, "D4_Accident"),
            "water_quality": yes_count(exp, "D5_Water_Quality"),
            "agri_livestock_affected": yes_count(exp, "D6_Agri Livestock Affected"),
            "displaced": yes_count(exp, "D7_Displaced\\Relocated"),
        },
        "sensitivity": {
            "land": freq(sens, "E1_ Land ownership"),
            "livestock": yes_count(sens, "E2_Does the household own livestock"),
            "safe_water": freq(sens, "E3_ Access to safe drinking water"),
            "toilet": freq(sens, "E4_ Access to toilet facility"),
            "debt": yes_count(sens, "E5_Debt"),
            "loan_source": freq(sens, "E5b_ Main source of loan"),
            "marginalization": yes_count(sens, "E6_Social marginalization"),
            "dependents": yes_count(sens, "E7. Presence of dependents"),
        },
        "adaptive_capacity": {
            "savings": yes_count(adap, "F1_Savings"),
            "bank": yes_count(adap, "F2_Bank_Account"),
            "group_member": yes_count(adap, "F3_Group_Member(SHG,Union)"),
            "social_protection": freq(adap, "F4_Access to government social protection"),
            "mining_compensation": yes_count(adap, "F5_Mining_Compensation"),
            "alt_livelihood": yes_count(adap, "F6_Alt_Livelihood"),
            "health_facility": freq(adap, "F7_Health_Facility"),
            "primary_school": yes_count(adap, "F8_Access to primary school"),
        },
        "potential_index_sheets": [],
    }

    for name, obj in sheets.items():
        headers = obj["headers"]
        if maybe_index_sheet(headers) or maybe_index_sheet([name]):
            compact = {"name": name, "headers": headers, "numeric_stats": {}}
            for header in headers:
                st = stats(obj["data"], header)
                if st and st["count"] >= 10:
                    compact["numeric_stats"][header] = st
            compact["sample_rows"] = obj["data"][:3]
            summary["potential_index_sheets"].append(compact)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
